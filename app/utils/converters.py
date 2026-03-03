import pandas as pd
import os
from io import BytesIO
from copy import copy, deepcopy
import zipfile
import json
from PIL import Image
import fitz  # PyMuPDF

# Mapeamento de extensões para formatos Excel
EXCEL_FORMATS = {
    '.xlsx': {'name': 'Excel 2007+ (.xlsx)', 'engine_read': 'openpyxl', 'engine_write': 'openpyxl'},
    '.xls': {'name': 'Excel 97-2003 (.xls)', 'engine_read': 'xlrd', 'engine_write': 'xlwt'},
    '.xlsb': {'name': 'Excel Binary (.xlsb)', 'engine_read': 'pyxlsb', 'engine_write': None},
    '.ods': {'name': 'OpenDocument (.ods)', 'engine_read': 'odf', 'engine_write': 'odf'},
}

MIMETYPES = {
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.xls': 'application/vnd.ms-excel',
    '.xlsb': 'application/vnd.ms-excel.sheet.binary.macroEnabled.12',
    '.ods': 'application/vnd.oasis.opendocument.spreadsheet',
    '.csv': 'text/csv',
    '.json': 'application/json',
    '.zip': 'application/zip',
    '.pdf': 'application/pdf',
    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
}

# Formatos de imagem suportados
SUPPORTED_IMAGE_FORMATS = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp']

def detect_excel_version(filename):
    """Detecta a versão do Excel baseado na extensão do arquivo."""
    ext = os.path.splitext(filename)[1].lower()
    if ext in EXCEL_FORMATS:
        return {
            'extension': ext,
            'format': EXCEL_FORMATS[ext]['name'],
            'available_conversions': get_available_conversions(ext)
        }
    return None

def get_available_conversions(current_ext):
    """Retorna as opções de conversão disponíveis para o formato atual."""
    conversions = []
    
    # Conversões para outros formatos Excel
    for ext, info in EXCEL_FORMATS.items():
        if ext != current_ext and info['engine_write'] is not None:
            conversions.append({
                'value': f'excel_to_{ext[1:]}',
                'label': f'Converter para {info["name"]}'
            })
    
    # Conversão para CSV (todas as planilhas em ZIP)
    conversions.append({
        'value': 'excel_to_csv',
        'label': 'Converter para CSV (ZIP com todas as planilhas)'
    })
    
    # Conversão para JSON (todas as planilhas)
    conversions.append({
        'value': 'excel_to_json',
        'label': 'Converter para JSON (todas as planilhas)'
    })
    
    return conversions


def convert_xlsx_preserving_all(source_file):
    """
    Converte xlsx preservando TODA a formatação.
    Carrega e salva diretamente o workbook - isso preserva:
    - Formatação condicional
    - Estilos de célula
    - Validação de dados
    - Filtros
    - Células mescladas
    - Comentários
    - Imagens
    - Etc.
    """
    from openpyxl import load_workbook
    
    output = BytesIO()
    
    # Carregar workbook preservando TUDO
    wb = load_workbook(
        source_file, 
        rich_text=True,
        data_only=False,
        keep_vba=False,
        keep_links=True
    )
    
    # Salvar diretamente - preserva toda a formatação original
    wb.save(output)
    wb.close()
    
    return output


def convert_xlsx_to_xls_all_sheets(source_file):
    """
    Converte xlsx para xls preservando o máximo possível de formatação.
    Nota: xls (Excel 97-2003) tem limitações inerentes ao formato.
    """
    import xlwt
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    
    output = BytesIO()
    source_wb = load_workbook(source_file, data_only=False, rich_text=True)
    target_wb = xlwt.Workbook(encoding='utf-8')
    
    # Mapeamento de cores RGB para índices xlwt
    def get_xlwt_color_index(rgb_str):
        """Converte RGB string para índice de cor xlwt aproximado."""
        if not rgb_str or len(rgb_str) < 6:
            return 0x7FFF
        rgb = rgb_str[-6:]
        try:
            r = int(rgb[0:2], 16)
            g = int(rgb[2:4], 16)
            b = int(rgb[4:6], 16)
        except:
            return 0x7FFF
        
        # Cores básicas xlwt
        colors = [
            ((0, 0, 0), 0),        # Preto
            ((255, 255, 255), 1),  # Branco
            ((255, 0, 0), 2),      # Vermelho
            ((0, 255, 0), 3),      # Verde
            ((0, 0, 255), 4),      # Azul
            ((255, 255, 0), 5),    # Amarelo
            ((255, 0, 255), 6),    # Magenta
            ((0, 255, 255), 7),    # Ciano
            ((128, 128, 128), 22), # Cinza
            ((192, 192, 192), 23), # Prata
        ]
        
        min_dist = float('inf')
        closest = 0
        for (cr, cg, cb), idx in colors:
            dist = (r-cr)**2 + (g-cg)**2 + (b-cb)**2
            if dist < min_dist:
                min_dist = dist
                closest = idx
        return closest
    
    def create_xlwt_style(cell):
        """Cria estilo xlwt baseado no estilo openpyxl."""
        style = xlwt.XFStyle()
        
        # Fonte
        font = xlwt.Font()
        if cell.font:
            font.name = cell.font.name or 'Arial'
            font.bold = cell.font.bold or False
            font.italic = cell.font.italic or False
            font.underline = cell.font.underline is not None
            if cell.font.size:
                font.height = int(cell.font.size * 20)
            if cell.font.color and cell.font.color.rgb and cell.font.color.rgb != '00000000':
                try:
                    font.colour_index = get_xlwt_color_index(cell.font.color.rgb)
                except:
                    pass
        style.font = font
        
        # Alinhamento
        alignment = xlwt.Alignment()
        if cell.alignment:
            h_align_map = {'left': 1, 'center': 2, 'right': 3, 'justify': 5, 'general': 0}
            v_align_map = {'top': 0, 'center': 1, 'bottom': 2, 'justify': 3}
            alignment.horz = h_align_map.get(cell.alignment.horizontal, 0)
            alignment.vert = v_align_map.get(cell.alignment.vertical, 1)
            alignment.wrap = 1 if cell.alignment.wrap_text else 0
        style.alignment = alignment
        
        # Padrão de preenchimento
        pattern = xlwt.Pattern()
        if cell.fill and cell.fill.fill_type == 'solid':
            pattern.pattern = xlwt.Pattern.SOLID_PATTERN
            if cell.fill.fgColor and cell.fill.fgColor.rgb:
                try:
                    pattern.pattern_fore_colour = get_xlwt_color_index(cell.fill.fgColor.rgb)
                except:
                    pass
        style.pattern = pattern
        
        # Bordas
        borders = xlwt.Borders()
        if cell.border:
            border_style_map = {
                'thin': 1, 'medium': 2, 'thick': 5, 'double': 6,
                'hair': 7, 'dashed': 3, 'dotted': 4, None: 0
            }
            if cell.border.left and cell.border.left.style:
                borders.left = border_style_map.get(cell.border.left.style, 0)
            if cell.border.right and cell.border.right.style:
                borders.right = border_style_map.get(cell.border.right.style, 0)
            if cell.border.top and cell.border.top.style:
                borders.top = border_style_map.get(cell.border.top.style, 0)
            if cell.border.bottom and cell.border.bottom.style:
                borders.bottom = border_style_map.get(cell.border.bottom.style, 0)
        style.borders = borders
        
        # Formato numérico
        if cell.number_format and cell.number_format != 'General':
            style.num_format_str = cell.number_format
        
        return style
    
    for sheet_name in source_wb.sheetnames:
        source_ws = source_wb[sheet_name]
        # xls tem limite de 31 caracteres no nome da sheet
        safe_name = sheet_name[:31]
        target_ws = target_wb.add_sheet(safe_name)
        
        # Copiar largura das colunas
        for col_letter, col_dim in source_ws.column_dimensions.items():
            if col_dim.width:
                col_idx = column_index_from_string(col_letter) - 1
                target_ws.col(col_idx).width = int(col_dim.width * 256)
        
        # Rastrear células mescladas para evitar duplicar escrita
        merged_cells = set()
        for merged_range in source_ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if row != merged_range.min_row or col != merged_range.min_col:
                        merged_cells.add((row - 1, col - 1))
            # Escrever merge
            first_cell = source_ws.cell(merged_range.min_row, merged_range.min_col)
            value = first_cell.value
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            try:
                style = create_xlwt_style(first_cell)
                target_ws.write_merge(
                    merged_range.min_row - 1, merged_range.max_row - 1,
                    merged_range.min_col - 1, merged_range.max_col - 1,
                    value, style
                )
            except:
                target_ws.write_merge(
                    merged_range.min_row - 1, merged_range.max_row - 1,
                    merged_range.min_col - 1, merged_range.max_col - 1,
                    str(value) if value else ''
                )
        
        # Copiar células com formatação
        for row_idx, row in enumerate(source_ws.iter_rows()):
            for col_idx, cell in enumerate(row):
                # Pular células que fazem parte de um merge (exceto a primeira)
                if (row_idx, col_idx) in merged_cells:
                    continue
                
                value = cell.value
                if value is not None:
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    
                    try:
                        style = create_xlwt_style(cell)
                        target_ws.write(row_idx, col_idx, value, style)
                    except:
                        try:
                            target_ws.write(row_idx, col_idx, str(value))
                        except:
                            pass
                elif cell.has_style:
                    try:
                        style = create_xlwt_style(cell)
                        target_ws.write(row_idx, col_idx, '', style)
                    except:
                        pass
        
        # Copiar altura das linhas
        for row_num, row_dim in source_ws.row_dimensions.items():
            if row_dim.height:
                target_ws.row(row_num - 1).height_mismatch = True
                target_ws.row(row_num - 1).height = int(row_dim.height * 20)
    
    source_wb.close()
    target_wb.save(output)
    return output


def convert_xlsx_to_ods_all_sheets(source_file):
    """Converte xlsx para ods preservando dados e estrutura."""
    from openpyxl import load_workbook
    
    output = BytesIO()
    source_wb = load_workbook(source_file, data_only=False, rich_text=True)
    
    with pd.ExcelWriter(output, engine='odf') as writer:
        for sheet_name in source_wb.sheetnames:
            source_ws = source_wb[sheet_name]
            
            # Extrair dados mantendo estrutura
            data = []
            max_col = source_ws.max_column or 1
            
            for row in source_ws.iter_rows(min_row=1, max_col=max_col):
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                data.append(row_data)
            
            if data and len(data) > 1:
                df = pd.DataFrame(data[1:], columns=data[0])
            elif data:
                df = pd.DataFrame([data[0]] if data else [])
            else:
                df = pd.DataFrame()
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    source_wb.close()
    return output


def convert_other_to_xlsx_all_sheets(source_file, source_ext):
    """Converte outros formatos para xlsx preservando todas as planilhas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output = BytesIO()
    engine = EXCEL_FORMATS.get(source_ext, {}).get('engine_read', 'openpyxl')
    
    excel_file = pd.ExcelFile(source_file, engine=engine)
    sheet_names = excel_file.sheet_names
    
    target_wb = Workbook()
    target_wb.remove(target_wb.active)
    
    for sheet_name in sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        target_ws = target_wb.create_sheet(title=sheet_name)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    value = None
                cell = target_ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
    
    target_wb.save(output)
    return output


def convert_excel_to_csv_all_sheets(source_file, source_ext, base_filename):
    """Converte todas as planilhas para CSV em um arquivo ZIP."""
    output = BytesIO()
    engine = EXCEL_FORMATS.get(source_ext, {}).get('engine_read', 'openpyxl')
    
    excel_file = pd.ExcelFile(source_file, engine=engine)
    sheet_names = excel_file.sheet_names
    
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for sheet_name in sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            csv_buffer = BytesIO()
            df.to_csv(csv_buffer, index=False, encoding='utf-8')
            safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
            zipf.writestr(f"{safe_name}.csv", csv_buffer.getvalue())
    
    return output


def convert_excel_to_json_all_sheets(source_file, source_ext):
    """Converte todas as planilhas para JSON."""
    output = BytesIO()
    engine = EXCEL_FORMATS.get(source_ext, {}).get('engine_read', 'openpyxl')
    
    excel_file = pd.ExcelFile(source_file, engine=engine)
    sheet_names = excel_file.sheet_names
    
    result = {}
    for sheet_name in sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        result[sheet_name] = json.loads(df.to_json(orient='records', date_format='iso'))
    
    json_str = json.dumps(result, indent=4, ensure_ascii=False, default=str)
    output.write(json_str.encode('utf-8'))
    return output


def read_excel_file(file, ext):
    """Lê arquivo Excel com o engine apropriado."""
    engine = EXCEL_FORMATS.get(ext, {}).get('engine_read', 'openpyxl')
    return pd.read_excel(file, engine=engine)


def convert_file(file, conv_type):
    output = BytesIO()
    base_filename = os.path.splitext(file.filename)[0]
    source_ext = os.path.splitext(file.filename)[1].lower()
    
    if conv_type == 'csv_to_xlsx':
        # dtype=str preserva zeros à esquerda em dados numéricos
        df = pd.read_csv(file, dtype=str)
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        output_filename = base_filename + '.xlsx'
        mimetype = MIMETYPES['.xlsx']
        
    elif conv_type == 'xlsx_to_csv':
        output = convert_excel_to_csv_all_sheets(file, '.xlsx', base_filename)
        output_filename = base_filename + '_csv.zip'
        mimetype = MIMETYPES['.zip']
        
    elif conv_type == 'csv_to_json':
        # dtype=str preserva zeros à esquerda em dados numéricos
        df = pd.read_csv(file, dtype=str)
        json_str = df.to_json(orient='records', indent=4)
        output.write(json_str.encode('utf-8'))
        output_filename = base_filename + '.json'
        mimetype = MIMETYPES['.json']
    
    # Conversões de Excel para xlsx - PRESERVA TODA FORMATAÇÃO
    elif conv_type == 'excel_to_xlsx':
        if source_ext == '.xlsx':
            # xlsx para xlsx - carrega e salva diretamente, preservando TUDO
            output = convert_xlsx_preserving_all(file)
        else:
            output = convert_other_to_xlsx_all_sheets(file, source_ext)
        output_filename = base_filename + '.xlsx'
        mimetype = MIMETYPES['.xlsx']
        
    elif conv_type == 'excel_to_xls':
        if source_ext == '.xlsx':
            output = convert_xlsx_to_xls_all_sheets(file)
        else:
            temp_output = convert_other_to_xlsx_all_sheets(file, source_ext)
            temp_output.seek(0)
            output = convert_xlsx_to_xls_all_sheets(temp_output)
        output_filename = base_filename + '.xls'
        mimetype = MIMETYPES['.xls']
        
    elif conv_type == 'excel_to_ods':
        if source_ext == '.xlsx':
            output = convert_xlsx_to_ods_all_sheets(file)
        else:
            temp_output = convert_other_to_xlsx_all_sheets(file, source_ext)
            temp_output.seek(0)
            output = convert_xlsx_to_ods_all_sheets(temp_output)
        output_filename = base_filename + '.ods'
        mimetype = MIMETYPES['.ods']
    
    elif conv_type == 'excel_to_csv':
        output = convert_excel_to_csv_all_sheets(file, source_ext, base_filename)
        output_filename = base_filename + '_csv.zip'
        mimetype = MIMETYPES['.zip']
    
    elif conv_type == 'excel_to_json':
        output = convert_excel_to_json_all_sheets(file, source_ext)
        output_filename = base_filename + '.json'
        mimetype = MIMETYPES['.json']
    
    elif conv_type == 'image_to_pdf':
        output = convert_image_to_pdf(file)
        output_filename = base_filename + '.pdf'
        mimetype = MIMETYPES['.pdf']
    
    elif conv_type == 'word_to_pdf':
        output = convert_word_to_pdf(file)
        output_filename = base_filename + '.pdf'
        mimetype = MIMETYPES['.pdf']
    
    elif conv_type == 'pdf_to_word':
        output = convert_pdf_to_word(file)
        output_filename = base_filename + '.docx'
        mimetype = MIMETYPES['.docx']
        
    else:
        raise ValueError("Tipo de conversão inválido")

    output.seek(0)
    return output, output_filename, mimetype


def convert_image_to_pdf(file):
    """
    Converte uma imagem para PDF.
    
    Args:
        file: Arquivo de imagem (FileStorage ou BytesIO)
        
    Returns:
        BytesIO contendo o PDF
    """
    # Ler a imagem
    img = Image.open(file)
    
    # Converter para RGB se necessário (ex: PNG com transparência)
    if img.mode in ('RGBA', 'LA', 'P'):
        # Criar background branco
        background = Image.new('RGB', img.size, (255, 255, 255))
        if img.mode == 'P':
            img = img.convert('RGBA')
        background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
        img = background
    elif img.mode != 'RGB':
        img = img.convert('RGB')
    
    # Salvar como PDF
    output = BytesIO()
    img.save(output, 'PDF', resolution=100.0)
    
    output.seek(0)
    return output


def convert_multiple_images_to_pdf(files):
    """
    Converte múltiplas imagens para um único PDF.
    
    Args:
        files: Lista de arquivos de imagem
        
    Returns:
        BytesIO contendo o PDF
    """
    images = []
    
    for file in files:
        img = Image.open(file)
        
        # Converter para RGB se necessário
        if img.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        images.append(img)
    
    if not images:
        raise ValueError("Nenhuma imagem válida encontrada")
    
    output = BytesIO()
    
    # Primeira imagem
    first_image = images[0]
    
    # Salvar todas as imagens no mesmo PDF
    if len(images) > 1:
        first_image.save(output, 'PDF', resolution=100.0, save_all=True, append_images=images[1:])
    else:
        first_image.save(output, 'PDF', resolution=100.0)
    
    output.seek(0)
    return output


def convert_word_to_pdf(file):
    """
    Converte um documento Word (.docx) para PDF.
    
    Args:
        file: Arquivo Word (FileStorage ou BytesIO)
        
    Returns:
        BytesIO contendo o PDF
    """
    import tempfile
    import os
    
    # Salvar arquivo temporário
    with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as tmp_docx:
        file.save(tmp_docx)
        tmp_docx_path = tmp_docx.name
    
    tmp_pdf_path = tmp_docx_path.replace('.docx', '.pdf')
    
    try:
        # Usar PyMuPDF para converter (método mais simples e confiável)
        # PyMuPDF pode abrir documentos do Word diretamente
        doc = fitz.open(tmp_docx_path)
        
        # Converter para PDF
        pdf_bytes = doc.convert_to_pdf()
        doc.close()
        
        output = BytesIO(pdf_bytes)
        output.seek(0)
        return output
        
    finally:
        # Limpar arquivos temporários
        if os.path.exists(tmp_docx_path):
            os.remove(tmp_docx_path)
        if os.path.exists(tmp_pdf_path):
            os.remove(tmp_pdf_path)


def convert_pdf_to_word(file):
    """
    Converte um documento PDF para Word (.docx).
    
    Args:
        file: Arquivo PDF (FileStorage ou BytesIO)
        
    Returns:
        BytesIO contendo o documento Word
    """
    from pdf2docx import Converter
    import tempfile
    import os
    
    # Salvar arquivo temporário
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
        file.save(tmp_pdf)
        tmp_pdf_path = tmp_pdf.name
    
    tmp_docx_path = tmp_pdf_path.replace('.pdf', '.docx')
    
    try:
        # Converter PDF para Word
        cv = Converter(tmp_pdf_path)
        cv.convert(tmp_docx_path)
        cv.close()
        
        # Ler o arquivo convertido
        output = BytesIO()
        with open(tmp_docx_path, 'rb') as f:
            output.write(f.read())
        
        output.seek(0)
        return output
        
    finally:
        # Limpar arquivos temporários
        if os.path.exists(tmp_pdf_path):
            os.remove(tmp_pdf_path)
        if os.path.exists(tmp_docx_path):
            os.remove(tmp_docx_path)
